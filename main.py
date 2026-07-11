import datetime
import os
import openpyxl
import colorama
from colorama import Fore, Back
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


green = PatternFill(fill_type="solid", fgColor="00FF00")
red = PatternFill(fill_type="solid", fgColor="FF0000")
yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")
관리자암호 = 'admin' #관리자 암호
위험연속결석 = 3
위험결석횟수 = 3

#  _____                                                                                                        _____ 
# ( ___ )                                                                                                      ( ___ )
#  |   |~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|   | 
#  |   |                                                                                                        |   | 
#  |   |  :'######:::'#######:::'#######::::::::'##:'##::::'##:'##::: ##::::::'##:::::'##:'########:'########:  |   | 
#  |   |  '##... ##:'##.... ##:'##.... ##::::::: ##: ##:::: ##: ###:: ##:::::: ##:'##: ##:... ##..:: ##.....::  |   | 
#  |   |   ##:::..::..::::: ##: ##:::: ##::::::: ##: ##:::: ##: ####: ##:::::: ##: ##: ##:::: ##:::: ##:::::::  |   | 
#  |   |  . ######:::'#######:: ##:::: ##::::::: ##: ##:::: ##: ## ## ##:::::: ##: ##: ##:::: ##:::: ######:::  |   | 
#  |   |  :..... ##::...... ##: ##:::: ##:'##::: ##: ##:::: ##: ##. ####:::::: ##: ##: ##:::: ##:::: ##...::::  |   | 
#  |   |  '##::: ##:'##:::: ##: ##:::: ##: ##::: ##: ##:::: ##: ##:. ###:'###: ##: ##: ##:::: ##:::: ##:::::::  |   | 
#  |   |  . ######::. #######::. #######::. ######::. #######:: ##::. ##: ###:. ###. ###::::: ##:::: ##:::::::  |   | 
#  |   |  :......::::.......::::.......::::......::::.......:::..::::..::...:::...::...::::::..:::::..::::::::  |   | 
#  |   |                                                                                                        |   | 
#  |___|~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~|___| 
# (_____)                                                                                                      (_____)

#만든사람: 이서준 (INSTA: @s3ojun.wtf / GITHUB: @s3ojun-wtf) 2026-07-06

#find Latest Date Column (마지막 날짜 열)
def findLDC(ws):
    날짜 = datetime.date.today().strftime('%F')

    마지막_날짜열 = 5
    for j in range(6, ws.max_column + 1):
        if str(ws.cell(row=1, column=j).value)[:10] == 날짜:
            마지막_날짜열 = j-1
            flag = False
            break
    
        elif str(ws.cell(row=1, column=j).value) is not None:
            마지막_날짜열 = j
    
    날짜_열 = 마지막_날짜열 + 1
    ws.cell(row=1, column=날짜_열).value = 날짜

    return ws, 날짜_열

#find Student Number
def findSTD_N(ws, 학번):
    r = -1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == int(학번):
            r = i
            break
    
    return r

#학생 연속 결석 횟수 & 위험 수준 처리 함수
def stdProcess(ws, stdRow, dateCol):
    absStreak = 0
    absCnt = 0
    preStreak = 0

    #연속 결석 & 결석횟수 체크
    for column in range(6, dateCol+1):
        if ws.cell(row=stdRow, column=column).value == "O":
            absStreak = 0
            preStreak+=1
            ws.cell(row=stdRow, column=column).fill = green
        else:
            ws.cell(row=stdRow, column=column).value = "X"
            ws.cell(row=stdRow, column=column).fill = red
            absStreak+=1
            absCnt+=1
            preStreak = 0
    
    #연속결석 기록
    ws.cell(row=stdRow, column=4).value = absStreak
    ws.cell(row=stdRow, column=5).value = absCnt #결석횟수 체크
    
    if absStreak >= 3:
        ws.cell(row=stdRow, column=3).value = "위험"
        ws.cell(row=stdRow, column=3).fill = red
    elif absStreak >=1 or absCnt >= 1:
        ws.cell(row=stdRow, column=3).value = "주의"
        ws.cell(row=stdRow, column=3).fill = yellow
    if  preStreak >= 3 or absCnt == 0:
        ws.cell(row=stdRow, column=3).value = "정상"
        ws.cell(row=stdRow, column=3).fill = green
    
    return ws

def main():
    wb = load_workbook("./data/database.xlsx")
    ws = wb["Sheet1"]
    학번 = input(Fore.CYAN + "학번: ")

    if not 학번.isdigit():
        print(Fore.RED + "[-] 입력이 잘못되었습니다.")
        main()
        return

    학번_행 = findSTD_N(ws, 학번)

    if 학번_행 == -1:
        print(Fore.RED + "[-] 당신은 석식 신청자가 아닙니다.")
        main()
        return
    
    (ws, 날짜_열) = findLDC(ws)

    if ws.cell(row=학번_행, column=날짜_열).value == "O": #오늘 날짜에 현재 학번이 석식을 이미 먹었는가
        print(Fore.RED + "[-] 당신은 이미 석식을 드셨습니다.")
        wb.save("./data/database.xlsx")
        main()
        return
    
    os.system(f"start ./학생/{학번}/사진.png") #학생 사진 띄우기
            
    tmp = input(Fore.CYAN + "사진이 일치합니까? (참: 1  / 거짓: 0) > ")
    if tmp == '0':
        print(Fore.RED + "[-] 사진이 일치하지 않습니다.")
        
    elif tmp == '1':
        ws.cell(row=학번_행, column=날짜_열).value = "O" #석식 처리
        ws.cell(row=학번_행, column=날짜_열).fill = green #석식 처리
        print(Fore.GREEN + "[+] 석식 등록 완료.")
    
    else:
        print(Fore.MAGENTA + "[?] 입력이 잘못됐습니다.")

    os.system('taskkill /f /fi "WINDOWTITLE eq 사진*"')
   
    (ws, 날짜_열) = findLDC(ws)
    ws = stdProcess(ws, 학번_행, 날짜_열)

    wb.save("./data/database.xlsx") #중간 저장
    main()
    return

if __name__ == "__main__":
    colorama.init()
    wb = load_workbook("./data/database.xlsx")
    ws = wb["Sheet1"]
    (ws, 날짜_열) = findLDC(ws)

    for i in range(2, ws.max_row + 1):
            ws = stdProcess(ws, i, 날짜_열)
    
    wb.save("./data/database.xlsx")
    main()